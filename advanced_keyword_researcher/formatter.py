"""Output formatting — json, markdown, excel."""

import io
import json
import math
import re

from scorer import format_number


# ── Response builder ──

def build_response(
    query: str,
    velocity_data: dict,
    opportunity_data: dict,
    channels: list[dict],
    trend_score: dict,
    opp_score: dict,
    views_median: dict,
    duration_analysis: dict,
    market_analysis: dict,
    topics_and_trends: dict | None,
    title_dna: dict | None = None,
    outlier_videos: list[dict] | None = None,
    engagement_data: list[dict] | None = None,
) -> dict:
    return {
        "query": query,
        "trendScore": trend_score,
        "opportunityScore": opp_score,
        "opportunityData": opportunity_data,
        "velocityData": velocity_data,
        "viewsMedian": views_median,
        "durationAnalysis": duration_analysis,
        "marketAnalysis": market_analysis,
        "channels": channels[:60],
        "topicsAndTrends": topics_and_trends,
        "titleDna": title_dna,
        "outlierVideos": outlier_videos,
        "engagementData": engagement_data,
    }


# ── JSON ──

def format_json(resp: dict) -> str:
    return json.dumps(resp, indent=2, ensure_ascii=False)


# ── Markdown ──

def format_markdown(resp: dict) -> str:
    lines: list[str] = []
    query = resp.get("query", "")
    lines.append(f"# Research Report: {query}")
    lines.append("")

    ts = resp.get("trendScore", {})
    os_ = resp.get("opportunityScore", {})
    od = resp.get("opportunityData", {})
    vm = resp.get("viewsMedian", {})
    da = resp.get("durationAnalysis", {})
    ma = resp.get("marketAnalysis", {})
    channels = resp.get("channels", [])
    tat = resp.get("topicsAndTrends", {})

    # Key Metrics
    lines.append("## Metricas Principais")
    lines.append("")
    lines.append("| Metrica | Valor |")
    lines.append("|---------|-------|")
    lines.append(f"| Trend Score | {ts.get('score', 0)}/100 ({ts.get('category', '')}) |")
    # Velocity rows
    velocities = ts.get("velocities", {})
    lines.append(f"| Views Per Day | {format_number(velocities.get('day', 0))} |")
    lines.append(f"| Views/Ano | {format_number(velocities.get('year', 0))} |")
    lines.append(f"| Views/Mes | {format_number(velocities.get('month', 0))} |")
    lines.append(f"| Views/Semana | {format_number(velocities.get('week', 0))} |")
    growth = ts.get("growth", {})
    w2y = growth.get("weekToYear", 0)
    lines.append(f"| Growth | {w2y}x |")
    lines.append(f"| Opportunity Score | {os_.get('score', 0)}/100 ({os_.get('category', '')}) |")
    lines.append(f"| View:Sub Outliers | {od.get('outlierPercent', 0)}% |")
    lines.append(f"| Competition | {ma.get('saturation', '')} ({format_number(ma.get('resultCount', 0))} resultados) |")
    lines.append("")

    # Channel Size Distribution
    metrics = od.get("metrics", {})
    lines.append("## Channel Size Distribution")
    lines.append("")
    lines.append("| Size | Count | Percent |")
    lines.append("|------|-------|---------|")
    for size in ["micro", "small", "medium", "large"]:
        m = metrics.get(size, {})
        label = {"micro": "Micro (<10K)", "small": "Small (10K-100K)", "medium": "Medium (100K-500K)", "large": "Large (500K+)"}[size]
        lines.append(f"| {label} | {m.get('count', 0)} | {m.get('percent', 0)}% |")
    lines.append(f"| **Outlier Videos** | | {od.get('outlierPercent', 0)}% |")
    lines.append(f"| **Small Channel Combined** | | {od.get('smallChannelPercent_combined', 0)}% |")
    lines.append("")

    # Duration Analysis
    lines.append("## Duration Analysis")
    lines.append("")
    lines.append("| Metrica | Valor |")
    lines.append("|---------|-------|")
    lines.append(f"| Mediana de duracao | {da.get('medianDuration', '')} |")
    lines.append(f"| Mediana de views (search + algo) | {format_number(vm.get('median', 0))} |")
    lines.append(f"| Q1 views | {format_number(vm.get('q1', 0))} |")
    lines.append(f"| Q3 views | {format_number(vm.get('q3', 0))} |")
    lines.append(f"| IQR (consistencia) | {format_number(vm.get('q3', 0) - vm.get('q1', 0))} |")
    lines.append("")

    # Estimated Audience Size
    if tat:
        lines.append("## Estimated Audience Size")
        lines.append("")
        lines.append(f"**{format_number(tat.get('estimatedAudienceSize', 0))}** {tat.get('audienceSizeExplanation', '')}")
        lines.append("")

    # Top Channels
    ch_limit = min(len(channels), 60)
    lines.append(f"## Top {ch_limit} Channels -- Ordem por Views Seen")
    lines.append("")
    lines.append("| # | Canal | Subs | Tamanho | Views Seen | Avg Views (Mean) | Avg Views (Median) | Ratio |")
    lines.append("|---|-------|------|---------|------------|-----------------|-------------------|-------|")
    for i, ch in enumerate(channels[:ch_limit], 1):
        cid = ch.get('channelId', '')
        title = ch.get('channelTitle', '')
        channel_link = f"[{title}](https://www.youtube.com/channel/{cid})" if cid else title
        subs = ch.get('subscriberCount', 0)
        size = ch.get('channelSize', '')
        size_label = {"micro": "Micro (<10K)", "small": "Small (10K-100K)", "medium": "Medium (100K-500K)", "large": "Large (500K+)"}.get(size, size)
        mean_str = format_number(ch.get('avgViewsPerVideoMean', 0))
        median_val = ch.get('avgViewsPerVideoMedian', 0)
        median_str = format_number(median_val) if median_val > 0 else "-"
        ratio = ch.get('meanMedianRatio', 0)
        ratio_str = f"{ratio}x" if ratio > 0 else "-"
        lines.append(f"| {i} | {channel_link} | {format_number(subs)} | {size_label} | {format_number(ch.get('viewsSeen', 0))} | {mean_str} | {median_str} | {ratio_str} |")
    lines.append("")

    # --- NEW SECTIONS ---

    # Top Videos by Outlier Score
    opp_videos_all = resp.get("outlierVideos", [])
    if opp_videos_all:
        # Sort by outlierMultiplier descending, filter out 0/inf/nan
        valid_videos = []
        for v in opp_videos_all:
            mult = v.get("outlierMultiplier", 0)
            if mult and mult > 0 and not (math.isinf(mult) or math.isnan(mult)):
                valid_videos.append(v)
        valid_videos.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

        if valid_videos:
            lines.append("### Top Videos por Outlier Score")
            lines.append("")
            lines.append("| # | Titulo | Canal | Views | Subs | Mult. | Dias |")
            lines.append("|---|--------|-------|-------|------|-------|------|")
            for i, v in enumerate(valid_videos, 1):
                vid = v.get("videoId", "")
                title = v.get("title", "")
                title_link = f"[{title}](https://www.youtube.com/watch?v={vid})" if vid else title
                ch_title = v.get("channelTitle", "")
                ch_info = v.get("channelInfo", {})
                subs = ch_info.get("subscriberCount", 0)
                mult = v.get("outlierMultiplier", 0)
                days = round(v.get("daysAgo", 0))
                lines.append(
                    f"| {i} | {title_link} | {ch_title} | "
                    f"{format_number(v.get('viewCount', 0))} | {format_number(subs)} | "
                    f"**{mult:.1f}x** | {days} |"
                )
            lines.append("")

    # Small Channel Outliers (<100K subs)
    if opp_videos_all:
        small_outliers = [
            v for v in opp_videos_all
            if v.get("channelSize") in ("micro", "small")
            and v.get("outlierMultiplier", 0) >= 1.0
            and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
        ]
        small_outliers.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

        if small_outliers:
            lines.append("### Outliers de Canais Pequenos (<100K inscritos)")
            lines.append("")
            lines.append("| # | Titulo | Canal | Views | Subs | Mult. | Dias | Tamanho |")
            lines.append("|---|--------|-------|-------|------|-------|------|---------|")
            for i, v in enumerate(small_outliers, 1):
                vid = v.get("videoId", "")
                title = v.get("title", "")
                title_link = f"[{title}](https://www.youtube.com/watch?v={vid})" if vid else title
                ch_title = v.get("channelTitle", "")
                ch_info = v.get("channelInfo", {})
                subs = ch_info.get("subscriberCount", 0)
                mult = v.get("outlierMultiplier", 0)
                days = round(v.get("daysAgo", 0))
                size = v.get("channelSize", "")
                size_label = "Micro" if size == "micro" else "Small"
                lines.append(
                    f"| {i} | {title_link} | {ch_title} | "
                    f"{format_number(v.get('viewCount', 0))} | {format_number(subs)} | "
                    f"**{mult:.1f}x** | {days} | {size_label} |"
                )
            lines.append("")

    # Outlier Videos (5x+)
    if opp_videos_all:
        outliers_5x = [
            v for v in opp_videos_all
            if v.get("outlierMultiplier", 0) >= 5.0
            and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
        ]
        outliers_5x.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

        lines.append("### Oportunidades Reais (videos 5x+ acima da media do canal)")
        lines.append("")

        if not outliers_5x:
            lines.append("Nenhum video com outlier 5x+ encontrado.")
            lines.append("")
        else:
            lines.append("| Titulo | Canal | Views | Mult. | Angulo provavel |")
            lines.append("|--------|-------|-------|-------|-----------------|")

            STOPWORDS_SHORT = frozenset({
                "the", "of", "to", "and", "in", "is", "it", "you", "that", "a", "an",
                "for", "on", "with", "as", "at", "by", "this", "be", "are", "was",
                "from", "or", "but", "not", "what", "all", "were", "we", "when",
                "your", "can", "said", "there", "use", "each", "which", "she",
                "do", "how", "their", "if", "will", "up", "about", "out", "many",
                "then", "them", "these", "so", "some", "her", "would", "make", "like",
                "him", "into", "time", "has", "look", "two", "more", "write", "go",
                "see", "no", "way", "could", "people", "my", "than", "first", "water",
                "been", "call", "who", "oil", "its", "now", "find", "long", "down",
                "day", "did", "get", "come", "made", "may", "part",
            })

            for v in outliers_5x:
                vid = v.get("videoId", "")
                title = v.get("title", "")
                title_link = f"[{title}](https://www.youtube.com/watch?v={vid})" if vid else title
                ch_title = v.get("channelTitle", "")
                mult = v.get("outlierMultiplier", 0)

                # Extract angle: non-stopword words + keyword
                words = re.findall(r"[a-zA-Z]+", title.lower())
                angle_words = [w for w in words if w not in STOPWORDS_SHORT and len(w) > 1]
                angle = " + ".join(angle_words[:5]) if angle_words else title[:40]

                lines.append(
                    f"| {title_link} | {ch_title} | "
                    f"{format_number(v.get('viewCount', 0))} | {mult:.1f}x | {angle} |"
                )
            lines.append("")

    # Title DNA
    title_dna = resp.get("titleDna")
    if title_dna:
        lines.append("### Titulo DNA -- Padroes dos Titulos")
        lines.append("")

        # Word frequency (top 10)
        word_freq = title_dna.get("wordFrequency", [])[:10]
        if word_freq:
            lines.append("| Palavra | Vezes | O que sugere |")
            lines.append("|---------|-------|---------------|")
            suggestions = {
                "psychology": "base do nicho (esperado)",
                "people": "foco em comportamento humano",
                "who": "formato de identificacao",
                "signs": "promessa de revelacao",
                "habits": "conteudo pratico",
                "men": "segmento popular",
                "women": "segmento popular",
                "gen x": "segmento popular",
                "gen z": "segmento popular",
                "millennial": "segmento popular",
                "actually": "subversao de expectativa",
                "secret": "promessa de exclusividade",
                "hidden": "promessa de revelacao",
                "sign": "promessa de revelacao",
                "habit": "conteudo pratico",
                "man": "segmento popular",
                "woman": "segmento popular",
                "why": "curiosidade",
                "how": "tutorial/educativo",
                "dont": "comportamento negativo",
                "don't": "comportamento negativo",
                "never": "intensidade/urgencia",
                "always": "generalizacao poderosa",
                "narcissist": "topicos de personalidade",
                "narcissists": "topicos de personalidade",
                "introvert": "segmento de personalidade",
                "introverts": "segmento de personalidade",
                "extrovert": "segmento de personalidade",
                "extroverts": "segmento de personalidade",
            }
            for wf in word_freq:
                word = wf.get("word", "")
                count = wf.get("count", 0)
                suggestion = suggestions.get(word.lower(), "")
                lines.append(f"| {word.capitalize()} | {count} | {suggestion} |")
            lines.append("")

        # Title structure patterns
        title_patterns = title_dna.get("titlePatterns", [])
        if title_patterns:
            lines.append("| Formato | Exemplo | Frequencia |")
            lines.append("|---------|---------|------------|")
            for tp in title_patterns:
                lines.append(
                    f"| {tp.get('format', '')} | {tp.get('example', '')[:50]} | "
                    f"{tp.get('frequency', '')} |"
                )
            lines.append("")

        # Enhanced section (only if LLM data exists)
        tat = resp.get("topicsAndTrends")
        if tat:
            lines.append("**Enhanced (LLM):**")
            lines.append("")
            # Show suggested topics if available
            suggested = tat.get("suggestedTopics", [])
            if suggested:
                lines.append("| Angulo | Potencial |")
                lines.append("|--------|-----------|")
                for t in suggested:
                    if isinstance(t, dict):
                        lines.append(f"| {t.get('topic', '')} | {t.get('opportunity', '')} |")
                lines.append("")

    # Engagement
    engagement_data = resp.get("engagementData", [])
    if engagement_data:
        lines.append("### Engajamento dos Top Videos")
        lines.append("")

        lines.append("| Titulo | Views | Likes | Eng. Rate | Sinal |")
        lines.append("|--------|-------|-------|-----------|-------|")
        for e in engagement_data:
            title = e.get("title", "")
            views = e.get("viewCount", 0)
            likes = e.get("likeCount", 0)
            eng_rate = e.get("engagementRate", 0)

            if likes == 0:
                eng_str = "N/A"
                signal = "N/A"
            else:
                eng_str = f"{eng_rate:.1f}%"
                if eng_rate < 2.0:
                    signal = "Fraco"
                elif eng_rate < 4.0:
                    signal = "Normal"
                elif eng_rate < 6.0:
                    signal = "Bom"
                else:
                    signal = "Excelente"

            lines.append(
                f"| {title} | {format_number(views)} | "
                f"{format_number(likes)} | {eng_str} | {signal} |"
            )
        lines.append("")

    # --- END NEW SECTIONS ---

    if tat:
        # Discovered Topics
        discovered = tat.get("discoveredTopics", [])
        if discovered:
            lines.append("## Discovered Topics")
            lines.append("")
            lines.append("| Topico | Videos | Avg Views |")
            lines.append("|--------|--------|-----------|")
            for t in discovered:
                if isinstance(t, dict):
                    lines.append(f"| {t.get('topic', '')} | {t.get('frequency', '')} | {format_number(t.get('avgViews', 0))} |")
                else:
                    lines.append(f"| {t} | - | - |")
            lines.append("")

        # Suggested Content Angles
        suggested = tat.get("suggestedTopics", [])
        if suggested:
            lines.append("## Suggested Content Angles")
            lines.append("")
            lines.append("| Angulo | Potencial | Justificativa |")
            lines.append("|--------|-----------|---------------|")
            for t in suggested:
                if isinstance(t, dict):
                    lines.append(f"| {t.get('topic', '')} | {t.get('opportunity', '')} | {t.get('reasoning', '')} |")
                else:
                    lines.append(f"| {t} | - | - |")
            lines.append("")

        # Audience Intent
        intent = tat.get("audienceIntent", {})
        if intent:
            lines.append("## Audience Intent")
            lines.append("")
            if isinstance(intent, dict):
                lines.append(f"**Primary:** {intent.get('primaryIntent', '')}")
                lines.append("")
                secondary = intent.get("secondaryIntents", [])
                if secondary:
                    lines.append("**Secondary:**")
                    for s in secondary:
                        lines.append(f"- {s}")
                    lines.append("")
                pain = intent.get("painPoints", [])
                if pain:
                    lines.append("**Pain Points:**")
                    for p in pain:
                        lines.append(f"- {p}")
                    lines.append("")
                motiv = intent.get("motivations", [])
                if motiv:
                    lines.append("**Motivations:**")
                    for m in motiv:
                        lines.append(f"- {m}")
                    lines.append("")
            else:
                lines.append(str(intent))
                lines.append("")

        # Psychographics
        psycho = tat.get("psychographics", {})
        if psycho:
            lines.append("## Audience Psychographics")
            lines.append("")
            if isinstance(psycho, dict):
                interests = psycho.get("interests", [])
                if interests:
                    lines.append(f"**Interesses:** {', '.join(interests)}")
                    lines.append("")
                values = psycho.get("values", [])
                if values:
                    lines.append(f"**Values:** {', '.join(values)}")
                    lines.append("")
                prefs = psycho.get("contentPreferences", [])
                if prefs:
                    lines.append("**Content Preferences:**")
                    for p in prefs:
                        lines.append(f"- {p}")
                    lines.append("")
            else:
                for p in (psycho if isinstance(psycho, list) else [psycho]):
                    lines.append(f"- {p}")
                lines.append("")

        # Key Insights
        summary = tat.get("summary", "")
        if summary:
            lines.append("## Key Insights")
            lines.append("")
            lines.append(summary)
            lines.append("")

    return "\n".join(lines)


# ── Excel ──

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

ENG_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
ENG_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
ENG_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

WRAP_COLS = {"Title", "Channel", "Example", "Insights", "Angle", "Reasoning", "Content Preference", "Primary Intent", "Secondary Intent", "Pain Point", "Motivation", "Interests", "Values", "Category", "Insight", "Topic", "Suggestion"}


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN


def _auto_width(ws, min_width=12, max_width=50, num_min=15):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        col_idx = col_cells[0].column
        header_val = ws.cell(row=1, column=col_idx).value
        header_len = len(str(header_val)) if header_val else 0

        is_num_col = False
        best = header_len
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                length = len(val_str)
                if length > best:
                    best = length
                if isinstance(cell.value, (int, float)):
                    is_num_col = True

        effective_min = num_min if is_num_col else min_width
        width = max(effective_min, best + 2)
        width = min(width, max_width)
        ws.column_dimensions[col_letter].width = width

        if header_val and header_val in WRAP_COLS:
            ws.column_dimensions[col_letter].width = max(width, 30)
            for cell in col_cells:
                cell.alignment = WRAP_ALIGN


def _fmt_int(n):
    if n is None:
        return ""
    n = int(n)
    if n == 0:
        return "-"
    if n >= 1_000_000:
        return f"{n:,.0f}"
    return f"{n:,}"


def _fmt_number_raw(n):
    if n is None or n == 0:
        return "-"
    return format_number(int(n))


def _fmt_mult(cell, mult):
    cell.value = f"{mult:.1f}x"
    if mult >= 5.0:
        cell.font = BOLD_FONT


def _add_hyperlink(ws, row, col, text, url):
    cell = ws.cell(row=row, column=col, value=text)
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    return cell


def _eng_fill(rate):
    if rate < 2.0:
        return ENG_RED
    if rate < 4.0:
        return ENG_YELLOW
    return ENG_GREEN


def _eng_signal(rate):
    if rate < 2.0:
        return "Fraco"
    if rate < 4.0:
        return "Normal"
    if rate < 6.0:
        return "Bom"
    return "Excelente"


def _write_overview(ws, resp):
    ws.append(["Metric", "Value"])
    ts = resp.get("trendScore", {})
    os_ = resp.get("opportunityScore", {})
    od = resp.get("opportunityData", {})
    vm = resp.get("viewsMedian", {})
    ma = resp.get("marketAnalysis", {})
    velocities = ts.get("velocities", {})
    growth = ts.get("growth", {})

    rows = [
        ("Trend Score", f"{ts.get('score', 0)}/100 ({ts.get('category', '')})"),
        ("Opportunity Score", f"{os_.get('score', 0)}/100 ({os_.get('category', '')})"),
        ("Views/Day", _fmt_number_raw(velocities.get("day", 0))),
        ("Views/Month", _fmt_number_raw(velocities.get("month", 0))),
        ("Views/Week", _fmt_number_raw(velocities.get("week", 0))),
        ("Views/Year", _fmt_number_raw(velocities.get("year", 0))),
        ("Growth", f"{growth.get('weekToYear', 0)}x"),
        ("Competition", f"{ma.get('saturation', '')} ({_fmt_int(ma.get('resultCount', 0))} results)"),
        ("Outlier %", f"{od.get('outlierPercent', 0)}%"),
        ("Saturation", ma.get("saturation", "")),
    ]
    for label, val in rows:
        ws.append([label, val])
    _style_header(ws, 1, 2)
    _auto_width(ws)


def _write_channels(ws, channels):
    ws.append(["#", "Channel", "Subscribers", "Channel Size", "Views Seen", "Avg Views (Mean)", "Avg Views (Median)", "Mean/Median Ratio"])
    for i, ch in enumerate(channels or [], 1):
        cid = ch.get("channelId", "")
        title = ch.get("channelTitle", "")
        url = f"https://www.youtube.com/channel/{cid}" if cid else ""
        subs = ch.get("subscriberCount", 0)
        size = ch.get("channelSize", "")
        size_label = {"micro": "Micro (<10K)", "small": "Small (10K-100K)", "medium": "Medium (100K-500K)", "large": "Large (500K+)"}.get(size, size)
        mean_val = ch.get("avgViewsPerVideoMean", 0)
        median_val = ch.get("avgViewsPerVideoMedian", 0)
        ratio = ch.get("meanMedianRatio", 0)
        ws.append([i, title,
                    _fmt_int(subs),
                    size_label,
                    _fmt_int(ch.get("viewsSeen", 0)),
                    _fmt_int(mean_val),
                    _fmt_int(median_val),
                    f"{ratio}x" if ratio > 0 else "-"])
        if url:
            _add_hyperlink(ws, i + 1, 2, title, url)
    _style_header(ws, 1, 8)
    _auto_width(ws)


def _write_outliers(ws, outlier_videos):
    ws.append(["#", "Title", "Channel", "Views", "Subs", "Multiplier", "Days", "URL"])
    valid = []
    for v in (outlier_videos or []):
        mult = v.get("outlierMultiplier", 0)
        if mult and mult > 0 and not (math.isinf(mult) or math.isnan(mult)):
            valid.append(v)
    valid.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    for i, v in enumerate(valid, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        subs = v.get("channelInfo", {}).get("subscriberCount", 0)
        row_num = i + 1
        ws.append([i, v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), _fmt_int(subs),
                    f"{mult:.1f}x", round(v.get("daysAgo", 0)), url])
        _add_hyperlink(ws, row_num, 2, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 8, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=6), mult)
    _style_header(ws, 1, 8)
    _auto_width(ws)


def _write_small_outliers(ws, outlier_videos):
    ws.append(["#", "Title", "Channel", "Views", "Subs", "Multiplier", "Days", "Size", "URL"])
    small = [
        v for v in (outlier_videos or [])
        if v.get("channelSize") in ("micro", "small")
        and v.get("outlierMultiplier", 0) >= 1.0
        and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
    ]
    small.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    for i, v in enumerate(small, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        subs = v.get("channelInfo", {}).get("subscriberCount", 0)
        size = v.get("channelSize", "")
        size_label = "Micro" if size == "micro" else "Small"
        row_num = i + 1
        ws.append([i, v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), _fmt_int(subs),
                    f"{mult:.1f}x", round(v.get("daysAgo", 0)), size_label, url])
        _add_hyperlink(ws, row_num, 2, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 9, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=6), mult)
    _style_header(ws, 1, 9)
    _auto_width(ws)


def _write_opportunities(ws, outlier_videos):
    ws.append(["Title", "Channel", "Views", "Multiplier", "Angle", "URL"])
    outliers_5x = [
        v for v in (outlier_videos or [])
        if v.get("outlierMultiplier", 0) >= 5.0
        and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
    ]
    outliers_5x.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    STOPWORDS = frozenset({
        "the", "of", "to", "and", "in", "is", "it", "you", "that", "a", "an",
        "for", "on", "with", "as", "at", "by", "this", "be", "are", "was",
        "from", "or", "but", "not", "what", "all", "were", "we", "when",
        "your", "can", "said", "there", "use", "each", "which", "she",
        "do", "how", "their", "if", "will", "up", "about", "out", "many",
        "then", "them", "these", "so", "some", "her", "would", "make", "like",
        "him", "into", "time", "has", "look", "two", "more", "write", "go",
        "see", "no", "way", "could", "people", "my", "than", "first", "water",
        "been", "call", "who", "oil", "its", "now", "find", "long", "down",
        "day", "did", "get", "come", "made", "may", "part",
    })

    for i, v in enumerate(outliers_5x, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        words = re.findall(r"[a-zA-Z]+", v.get("title", "").lower())
        angle_words = [w for w in words if w not in STOPWORDS and len(w) > 1]
        angle = " + ".join(angle_words[:5]) if angle_words else v.get("title", "")[:40]
        row_num = i + 1
        ws.append([v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), f"{mult:.1f}x", angle, url])
        _add_hyperlink(ws, row_num, 1, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 6, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=4), mult)
    _style_header(ws, 1, 6)
    _auto_width(ws)


def _write_title_dna_words(ws, title_dna):
    ws.append(["Word", "Count", "Suggestion"])
    word_freq = title_dna.get("wordFrequency", [])[:10]
    suggestions = {
        "psychology": "base do nicho (esperado)",
        "people": "foco em comportamento humano",
        "who": "formato de identificacao",
        "signs": "promessa de revelacao",
        "habits": "conteudo pratico",
        "men": "segmento popular",
        "women": "segmento popular",
        "gen x": "segmento popular",
        "gen z": "segmento popular",
        "millennial": "segmento popular",
        "actually": "subversao de expectativa",
        "secret": "promessa de exclusividade",
        "hidden": "promessa de revelacao",
        "sign": "promessa de revelacao",
        "habit": "conteudo pratico",
        "man": "segmento popular",
        "woman": "segmento popular",
        "why": "curiosidade",
        "how": "tutorial/educativo",
        "dont": "comportamento negativo",
        "don't": "comportamento negativo",
        "never": "intensidade/urgencia",
        "always": "generalizacao poderosa",
        "narcissist": "topicos de personalidade",
        "narcissists": "topicos de personalidade",
        "introvert": "segmento de personalidade",
        "introverts": "segmento de personalidade",
        "extrovert": "segmento de personalidade",
        "extroverts": "segmento de personalidade",
    }
    for wf in word_freq:
        word = wf.get("word", "")
        count = wf.get("count", 0)
        suggestion = suggestions.get(word.lower(), "")
        ws.append([word.capitalize(), count, suggestion])
    _style_header(ws, 1, 3)
    _auto_width(ws)


def _write_title_dna_patterns(ws, title_dna):
    ws.append(["Format", "Example", "Frequency"])
    title_patterns = title_dna.get("titlePatterns", [])
    for tp in title_patterns:
        ws.append([tp.get("format", ""), tp.get("example", ""), tp.get("frequency", "")])
    _style_header(ws, 1, 3)
    ws.column_dimensions["B"].width = 60
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_ALIGN
    _auto_width(ws)


def _write_engagement(ws, engagement_data):
    ws.append(["Title", "Views", "Likes", "Eng. Rate", "Signal"])
    for i, e in enumerate(engagement_data or [], 1):
        views = e.get("viewCount", 0)
        likes = e.get("likeCount", 0)
        eng_rate = e.get("engagementRate", 0)
        if likes == 0:
            eng_str = "N/A"
            signal = "N/A"
        else:
            eng_str = f"{eng_rate:.1f}%"
            signal = _eng_signal(eng_rate)
        row_num = i + 1
        ws.append([e.get("title", ""), _fmt_int(views), _fmt_int(likes), eng_str, signal])
        if likes > 0:
            ws.cell(row=row_num, column=4).fill = _eng_fill(eng_rate)
    _style_header(ws, 1, 5)
    _auto_width(ws)


def _write_audience(ws, resp):
    ws.append(["Category", "Insight"])
    tat = resp.get("topicsAndTrends") or {}
    intent = tat.get("audienceIntent", {})
    psycho = tat.get("psychographics", {})
    row = 2

    if isinstance(intent, dict):
        primary = intent.get("primaryIntent", "")
        if primary:
            ws.cell(row=row, column=1, value="Primary Intent")
            ws.cell(row=row, column=2, value=primary)
            row += 1

        for s in intent.get("secondaryIntents", []):
            ws.cell(row=row, column=1, value="Secondary Intent")
            ws.cell(row=row, column=2, value=s)
            row += 1

        for p in intent.get("painPoints", []):
            ws.cell(row=row, column=1, value="Pain Point")
            ws.cell(row=row, column=2, value=p)
            row += 1

        for m in intent.get("motivations", []):
            ws.cell(row=row, column=1, value="Motivation")
            ws.cell(row=row, column=2, value=m)
            row += 1

    if isinstance(psycho, dict):
        interests = psycho.get("interests", [])
        if interests:
            ws.cell(row=row, column=1, value="Interests")
            ws.cell(row=row, column=2, value=", ".join(interests))
            row += 1

        values = psycho.get("values", [])
        if values:
            ws.cell(row=row, column=1, value="Values")
            ws.cell(row=row, column=2, value=", ".join(values))
            row += 1

        for p in psycho.get("contentPreferences", []):
            ws.cell(row=row, column=1, value="Content Preference")
            ws.cell(row=row, column=2, value=p)
            row += 1

    _style_header(ws, 1, 2)
    _auto_width(ws)


def _write_topics(ws, resp):
    ws.append(["Topic", "Videos", "Avg Views", "Angle", "Potential", "Reasoning"])
    tat = resp.get("topicsAndTrends") or {}
    row = 2

    discovered = tat.get("discoveredTopics", [])
    for t in discovered:
        if isinstance(t, dict):
            ws.cell(row=row, column=1, value=t.get("topic", ""))
            ws.cell(row=row, column=2, value=t.get("frequency", ""))
            ws.cell(row=row, column=3, value=_fmt_int(t.get("avgViews", 0)))
            row += 1

    suggested = tat.get("suggestedTopics", [])
    for t in suggested:
        if isinstance(t, dict):
            ws.cell(row=row, column=1, value=t.get("topic", ""))
            ws.cell(row=row, column=4, value=t.get("topic", ""))
            ws.cell(row=row, column=5, value=t.get("opportunity", ""))
            ws.cell(row=row, column=6, value=t.get("reasoning", ""))
            row += 1

    _style_header(ws, 1, 6)
    _auto_width(ws)


def _write_insights(ws, resp):
    ws.append(["Insights"])
    tat = resp.get("topicsAndTrends") or {}
    summary = tat.get("summary", "")
    if summary:
        ws.cell(row=2, column=1, value=summary).alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 120
    _style_header(ws, 1, 1)


def format_excel(resp: dict) -> bytes:
    wb = Workbook()
    channels = resp.get("channels") or []
    outlier_videos = resp.get("outlierVideos") or []
    engagement_data = resp.get("engagementData") or []
    title_dna = resp.get("titleDna")

    writers = [
        ("Overview", lambda ws: _write_overview(ws, resp)),
        ("Channels", lambda ws: _write_channels(ws, channels)),
        ("Outliers", lambda ws: _write_outliers(ws, outlier_videos)),
        ("Small Outliers", lambda ws: _write_small_outliers(ws, outlier_videos)),
        ("Opportunities", lambda ws: _write_opportunities(ws, outlier_videos)),
        ("Title DNA - Words", lambda ws: _write_title_dna_words(ws, title_dna) if title_dna else ws.append(["No data"])),
        ("Title DNA - Patterns", lambda ws: _write_title_dna_patterns(ws, title_dna) if title_dna else ws.append(["No data"])),
        ("Engagement", lambda ws: _write_engagement(ws, engagement_data)),
        ("Audience", lambda ws: _write_audience(ws, resp)),
        ("Topics", lambda ws: _write_topics(ws, resp)),
        ("Insights", lambda ws: _write_insights(ws, resp)),
    ]

    first = True
    for name, writer in writers:
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        writer(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Router ──

def format_output(resp: dict, fmt: str = "markdown") -> str | bytes:
    """Roteador: retorna str (json/markdown) ou bytes (excel)."""
    match fmt:
        case "json":
            return format_json(resp)
        case "markdown":
            return format_markdown(resp)
        case "excel":
            return format_excel(resp)
        case _:
            raise ValueError(f"Unknown format: {fmt}")
