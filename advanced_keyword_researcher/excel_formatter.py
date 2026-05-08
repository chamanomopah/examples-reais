"""Excel output formatting."""

import io
import math
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from scorer import format_number

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

ENG_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
ENG_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
ENG_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN


def _auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_width
        for cell in col_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > best:
                    best = length
        ws.column_dimensions[col_letter].width = min(best + 2, max_width)


def _fmt_int(n):
    if n is None:
        return ""
    n = int(n)
    if n >= 1_000_000:
        return f"{n:,.0f}"
    return f"{n:,}" if n != 0 else "0"


def _fmt_number_raw(n):
    return format_number(int(n)) if n else "0"


def _fmt_mult(cell, mult):
    cell.value = f"{mult:.1f}x"
    if mult >= 5.0:
        cell.font = BOLD_FONT


def _add_hyperlink(ws, row, col, text, url):
    cell = ws.cell(row=row, column=col, value=text)
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    return cell


def _eng_fill(rate):
    if rate < 2.0:
        return ENG_RED
    if rate < 4.0:
        return ENG_YELLOW
    return ENG_GREEN


def _eng_signal(rate):
    if rate < 2.0:
        return "Fraco"
    if rate < 4.0:
        return "Normal"
    if rate < 6.0:
        return "Bom"
    return "Excelente"


def _write_overview(ws, resp):
    ws.append(["Metric", "Value"])
    ts = resp.get("trendScore", {})
    os_ = resp.get("opportunityScore", {})
    od = resp.get("opportunityData", {})
    vm = resp.get("viewsMedian", {})
    ma = resp.get("marketAnalysis", {})
    velocities = ts.get("velocities", {})
    growth = ts.get("growth", {})

    rows = [
        ("Trend Score", f"{ts.get('score', 0)}/100 ({ts.get('category', '')})"),
        ("Opportunity Score", f"{os_.get('score', 0)}/100 ({os_.get('category', '')})"),
        ("Views/Day", _fmt_number_raw(velocities.get("day", 0))),
        ("Views/Month", _fmt_number_raw(velocities.get("month", 0))),
        ("Views/Week", _fmt_number_raw(velocities.get("week", 0))),
        ("Views/Year", _fmt_number_raw(velocities.get("year", 0))),
        ("Growth", f"{growth.get('weekToYear', 0)}x"),
        ("Competition", f"{ma.get('saturation', '')} ({_fmt_int(ma.get('resultCount', 0))} results)"),
        ("Outlier %", f"{od.get('outlierPercent', 0)}%"),
        ("Saturation", ma.get("saturation", "")),
    ]
    for label, val in rows:
        ws.append([label, val])
    _style_header(ws, 1, 2)
    _auto_width(ws)


def _write_channels(ws, channels):
    ws.append(["#", "Channel", "Views Seen", "Avg Views (Mean)", "Avg Views (Median)"])
    for i, ch in enumerate(channels, 1):
        cid = ch.get("channelId", "")
        title = ch.get("channelTitle", "")
        url = f"https://www.youtube.com/channel/{cid}" if cid else ""
        ws.append([i, title, _fmt_int(ch.get("viewsSeen", 0)),
                    _fmt_int(ch.get("avgViewsPerVideoMean", 0)),
                    _fmt_int(ch.get("avgViewsPerVideoMedian", 0))])
        if url:
            _add_hyperlink(ws, i + 1, 2, title, url)
    _style_header(ws, 1, 5)
    _auto_width(ws)


def _write_outliers(ws, outlier_videos):
    ws.append(["#", "Title", "Channel", "Views", "Subs", "Multiplier", "Days", "URL"])
    valid = []
    for v in outlier_videos:
        mult = v.get("outlierMultiplier", 0)
        if mult and not (math.isinf(mult) or math.isnan(mult)):
            valid.append(v)
    valid.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    for i, v in enumerate(valid, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        subs = v.get("channelInfo", {}).get("subscriberCount", 0)
        row_num = i + 1
        ws.append([i, v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), _fmt_int(subs),
                    f"{mult:.1f}x", round(v.get("daysAgo", 0)), url])
        _add_hyperlink(ws, row_num, 2, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 8, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=6), mult)
    _style_header(ws, 1, 8)
    _auto_width(ws)


def _write_small_outliers(ws, outlier_videos):
    ws.append(["#", "Title", "Channel", "Views", "Subs", "Multiplier", "Days", "Size", "URL"])
    small = [
        v for v in outlier_videos
        if v.get("channelSize") in ("micro", "small")
        and v.get("outlierMultiplier", 0) >= 1.0
        and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
    ]
    small.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    for i, v in enumerate(small, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        subs = v.get("channelInfo", {}).get("subscriberCount", 0)
        size = v.get("channelSize", "")
        size_label = "Micro" if size == "micro" else "Small"
        row_num = i + 1
        ws.append([i, v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), _fmt_int(subs),
                    f"{mult:.1f}x", round(v.get("daysAgo", 0)), size_label, url])
        _add_hyperlink(ws, row_num, 2, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 9, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=6), mult)
    _style_header(ws, 1, 9)
    _auto_width(ws)


def _write_opportunities(ws, outlier_videos):
    ws.append(["Title", "Channel", "Views", "Multiplier", "Angle", "URL"])
    outliers_5x = [
        v for v in outlier_videos
        if v.get("outlierMultiplier", 0) >= 5.0
        and not (math.isinf(v.get("outlierMultiplier", 0)) or math.isnan(v.get("outlierMultiplier", 0)))
    ]
    outliers_5x.sort(key=lambda x: x.get("outlierMultiplier", 0), reverse=True)

    STOPWORDS = frozenset({
        "the", "of", "to", "and", "in", "is", "it", "you", "that", "a", "an",
        "for", "on", "with", "as", "at", "by", "this", "be", "are", "was",
        "from", "or", "but", "not", "what", "all", "were", "we", "when",
        "your", "can", "said", "there", "use", "each", "which", "she",
        "do", "how", "their", "if", "will", "up", "about", "out", "many",
        "then", "them", "these", "so", "some", "her", "would", "make", "like",
        "him", "into", "time", "has", "look", "two", "more", "write", "go",
        "see", "no", "way", "could", "people", "my", "than", "first", "water",
        "been", "call", "who", "oil", "its", "now", "find", "long", "down",
        "day", "did", "get", "come", "made", "may", "part",
    })

    for i, v in enumerate(outliers_5x, 1):
        vid = v.get("videoId", "")
        url = f"https://www.youtube.com/watch?v={vid}" if vid else ""
        mult = v.get("outlierMultiplier", 0)
        words = re.findall(r"[a-zA-Z]+", v.get("title", "").lower())
        angle_words = [w for w in words if w not in STOPWORDS and len(w) > 1]
        angle = " + ".join(angle_words[:5]) if angle_words else v.get("title", "")[:40]
        row_num = i + 1
        ws.append([v.get("title", ""), v.get("channelTitle", ""),
                    _fmt_int(v.get("viewCount", 0)), f"{mult:.1f}x", angle, url])
        _add_hyperlink(ws, row_num, 1, v.get("title", ""), url)
        _add_hyperlink(ws, row_num, 6, "Watch", url)
        _fmt_mult(ws.cell(row=row_num, column=4), mult)
    _style_header(ws, 1, 6)
    _auto_width(ws)


def _write_title_dna(ws, title_dna):
    ws.append(["Word", "Count", "Suggestion"])
    word_freq = title_dna.get("wordFrequency", [])[:10]
    suggestions = {
        "psychology": "base do nicho (esperado)",
        "people": "foco em comportamento humano",
        "who": "formato de identificacao",
        "signs": "promessa de revelacao",
        "habits": "conteudo pratico",
        "men": "segmento popular",
        "women": "segmento popular",
        "gen x": "segmento popular",
        "gen z": "segmento popular",
        "millennial": "segmento popular",
        "actually": "subversao de expectativa",
        "secret": "promessa de exclusividade",
        "hidden": "promessa de revelacao",
        "sign": "promessa de revelacao",
        "habit": "conteudo pratico",
        "man": "segmento popular",
        "woman": "segmento popular",
        "why": "curiosidade",
        "how": "tutorial/educativo",
        "dont": "comportamento negativo",
        "don't": "comportamento negativo",
        "never": "intensidade/urgencia",
        "always": "generalizacao poderosa",
        "narcissist": "topicos de personalidade",
        "narcissists": "topicos de personalidade",
        "introvert": "segmento de personalidade",
        "introverts": "segmento de personalidade",
        "extrovert": "segmento de personalidade",
        "extroverts": "segmento de personalidade",
    }
    for wf in word_freq:
        word = wf.get("word", "")
        count = wf.get("count", 0)
        suggestion = suggestions.get(word.lower(), "")
        ws.append([word.capitalize(), count, suggestion])

    title_patterns = title_dna.get("titlePatterns", [])
    if title_patterns:
        row = len(word_freq) + 3
        ws.cell(row=row, column=1, value="Title Patterns").font = BOLD_FONT
        row += 1
        ws.append([])  # blank row for spacing
        ws.append(["Format", "Example", "Frequency"])
        _style_header(ws, row + 1, 3)
        for tp in title_patterns:
            ws.append([tp.get("format", ""), tp.get("example", "")[:50], tp.get("frequency", "")])

    _style_header(ws, 1, 3)
    _auto_width(ws)


def _write_engagement(ws, engagement_data):
    ws.append(["Title", "Views", "Likes", "Eng. Rate", "Signal"])
    for i, e in enumerate(engagement_data, 1):
        views = e.get("viewCount", 0)
        likes = e.get("likeCount", 0)
        eng_rate = e.get("engagementRate", 0)
        if likes == 0:
            eng_str = "N/A"
            signal = "N/A"
        else:
            eng_str = f"{eng_rate:.1f}%"
            signal = _eng_signal(eng_rate)
        row_num = i + 1
        ws.append([e.get("title", ""), _fmt_int(views), _fmt_int(likes), eng_str, signal])
        if likes > 0:
            ws.cell(row=row_num, column=4).fill = _eng_fill(eng_rate)
    _style_header(ws, 1, 5)
    _auto_width(ws)


def _write_audience(ws, resp):
    ws.append(["Category", "Insight"])
    tat = resp.get("topicsAndTrends") or {}
    intent = tat.get("audienceIntent", {})
    psycho = tat.get("psychographics", {})
    row = 2

    if isinstance(intent, dict):
        primary = intent.get("primaryIntent", "")
        if primary:
            ws.cell(row=row, column=1, value="Primary Intent")
            ws.cell(row=row, column=2, value=primary)
            row += 1

        for s in intent.get("secondaryIntents", []):
            ws.cell(row=row, column=1, value="Secondary Intent")
            ws.cell(row=row, column=2, value=s)
            row += 1

        for p in intent.get("painPoints", []):
            ws.cell(row=row, column=1, value="Pain Point")
            ws.cell(row=row, column=2, value=p)
            row += 1

        for m in intent.get("motivations", []):
            ws.cell(row=row, column=1, value="Motivation")
            ws.cell(row=row, column=2, value=m)
            row += 1

    if isinstance(psycho, dict):
        interests = psycho.get("interests", [])
        if interests:
            ws.cell(row=row, column=1, value="Interests")
            ws.cell(row=row, column=2, value=", ".join(interests))
            row += 1

        values = psycho.get("values", [])
        if values:
            ws.cell(row=row, column=1, value="Values")
            ws.cell(row=row, column=2, value=", ".join(values))
            row += 1

        for p in psycho.get("contentPreferences", []):
            ws.cell(row=row, column=1, value="Content Preference")
            ws.cell(row=row, column=2, value=p)
            row += 1

    _style_header(ws, 1, 2)
    _auto_width(ws)


def _write_topics(ws, resp):
    ws.append(["Topic", "Videos", "Avg Views", "Angle", "Potential", "Reasoning"])
    tat = resp.get("topicsAndTrends") or {}
    row = 2

    discovered = tat.get("discoveredTopics", [])
    for t in discovered:
        if isinstance(t, dict):
            ws.cell(row=row, column=1, value=t.get("topic", ""))
            ws.cell(row=row, column=2, value=t.get("frequency", ""))
            ws.cell(row=row, column=3, value=_fmt_int(t.get("avgViews", 0)))
            row += 1

    suggested = tat.get("suggestedTopics", [])
    for t in suggested:
        if isinstance(t, dict):
            ws.cell(row=row, column=1, value=t.get("topic", ""))
            ws.cell(row=row, column=4, value=t.get("topic", ""))
            ws.cell(row=row, column=5, value=t.get("opportunity", ""))
            ws.cell(row=row, column=6, value=t.get("reasoning", ""))
            row += 1

    _style_header(ws, 1, 6)
    _auto_width(ws)


def _write_insights(ws, resp):
    ws.append(["Insights"])
    tat = resp.get("topicsAndTrends") or {}
    summary = tat.get("summary", "")
    ws.cell(row=2, column=1, value=summary).alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 120
    _style_header(ws, 1, 1)


def format_excel(resp: dict) -> bytes:
    wb = Workbook()
    channels = resp.get("channels", [])
    outlier_videos = resp.get("outlierVideos", [])
    engagement_data = resp.get("engagementData", [])
    title_dna = resp.get("titleDna")

    writers = [
        ("Overview", lambda ws: _write_overview(ws, resp)),
        ("Channels", lambda ws: _write_channels(ws, channels)),
        ("Outliers", lambda ws: _write_outliers(ws, outlier_videos)),
        ("Small Outliers", lambda ws: _write_small_outliers(ws, outlier_videos)),
        ("Opportunities", lambda ws: _write_opportunities(ws, outlier_videos)),
        ("Title DNA", lambda ws: _write_title_dna(ws, title_dna) if title_dna else None),
        ("Engagement", lambda ws: _write_engagement(ws, engagement_data)),
        ("Audience", lambda ws: _write_audience(ws, resp)),
        ("Topics", lambda ws: _write_topics(ws, resp)),
        ("Insights", lambda ws: _write_insights(ws, resp)),
    ]

    first = True
    for name, writer in writers:
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        result = writer(ws)
        if result is None:
            wb.remove(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
